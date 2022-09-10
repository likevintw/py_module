
# .xlsx
import unittest
import openpyxl

# python3 -m unittest -v


class TestFolderPathHandler(unittest.TestCase):
    @unittest.skip("Pass")
    def test_show_sheet_title(self):
        workbook = openpyxl.load_workbook("test.xlsx")
        for sheet in workbook:
            print(sheet.title)  # get a string

    @unittest.skip("Pass")
    def test_writr_and_read_cell_data(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        # write
        for sheet in workbook:
            sheet.cell(row=1, column=1, value=1)
            break
        # read
        data = None
        for sheet in workbook:
            data = sheet.cell(row=1, column=1).value
            break
        self.assertEqual(data, 1)

    @unittest.skip("Pass")
    def test_save_excel(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        # write
        for sheet in workbook:
            for i in range(10):
                sheet.cell(row=i+1, column=1, value=i)
            break
        # save
        workbook.save("demonstration.xlsx")

    @unittest.skip("Pass")
    def test_show_current_sheet(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        print(workbook.active)

    @unittest.skip("Pass")
    def test_change_sheet_label_color(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        for i in workbook:
            sheet = i
            break
        print(sheet.title)  # sheet name
        sheet.sheet_properties.tabColor = "1072BA"
        workbook.save("change_sheet_color.xlsx")

    @unittest.skip("Online Example")
    def test_create_sheet(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        # add one at first
        workbook.create_sheet("first_page", 0)
        # add one at latest
        workbook.create_sheet("lastest_page")

    @unittest.skip("Online Example")
    def test_copy_sheet(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        source = workbook.active
        workbook.copy_worksheet(source)

    @unittest.skip("Online Example")
    def test_show_worksheet_name(self):
        workbook = openpyxl.load_workbook("demonstration.xlsx")
        print(workbook.sheetnames)  # get a list


if __name__ == '__main__':
    unittest.main()
