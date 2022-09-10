
from cmath import e
from csv import excel
import string
import openpyxl

# workbook > sheet > cell


class ExcelHandler:
    def __init__(self) -> None:
        self.workbook = None

    def write_cell(self, sheet, row, column, value) -> string:
        if sheet == None:
            return "write_cell Error, sheet is empty"
        if row == None:
            return "write_cell Error, row is empty"
        if column == None:
            return "write_cell Error, column is empty"
        if value == None:
            return "write_cell Error, value is empty"

        try:
            sheet.cell(row=1, column=1, value=1)
        except e:
            return "write_cell Error, sheet.cell(row, column, value))".format(
                str(sheet), row, column, value)

    def show_current_sheet_name(self) -> string:
        try:
            return str(self.workbook.active())
        except:
            return "show_current_sheet_name Error"

    def change_sheet_label_color(
            self, color_code="1072BA",
            excel_path="export.xlsx") -> string:

        try:
            sheet = self.workbook[0]    # first sheet
        except:
            return "change_sheet_label_color, get sheet Error"

        print("change sheet label_color sheet name: ", sheet.title)

        try:
            sheet.sheet_properties.tabColor = color_code
        except:
            return "change_sheet_label_color, set sheet color Error"

        try:
            self.workbook.save(excel_path)
        except:
            return "change_sheet_label_color, export xlsx file Error"

        return None


def get_excel_handler(excel_path) -> ExcelHandler:
    handler = ExcelHandler()
    workbook = openpyxl.load_workbook(excel_path)
    handler.workbook = workbook
    return handler
